"""
Generate a disambiguation review workbook from analysis data.

Model-agnostic — accepts analysis results via JSON input and produces
disambiguation_results.xlsx with four sheets:
  1. Column Descriptions — proposed description text for each column
  2. Business Rules — defaults, canonical columns, routing logic
  3. Schema Exclusions — columns to hide from AI data schema
  4. Implementation Checklist — sequenced post-review actions

Each review row has Review Status / Reviewer / Comments for domain expert sign-off.
Nothing should be applied to the model until the relevant rows are Approved.
Once approved, the Implementation Checklist sheet defines the execution order.

Usage
-----
# From JSON input file:
python generate_disambiguation_excel.py --input analysis.json --output results.xlsx

# From stdin (piped from the agent):
cat analysis.json | python generate_disambiguation_excel.py --output results.xlsx

# The agent can also call generate_workbook() directly as a library:
from generate_disambiguation_excel import generate_workbook
generate_workbook(data, Path("results.xlsx"))

JSON input format
-----------------
{
  "model_name": "My_Semantic_Model",
  "workspace": "My_Workspace",
  "overlap_groups": [
    {
      "domain": "Country",
      "overlap_group": "Country",
      "ambiguity_class": "A – Competing Classifications",
      "when_user_says": "a country name (\"India\", \"US\") or \"by country\"",
      "default_column": "Dim_Customer.Ship-To Country",
      "priority": "High",
      "columns": [
        {
          "table": "Dim_Customer",
          "column": "Ship-To Country",
          "current_description": "...",
          "proposed_description": "...",
          "suggested_synonyms": "syn1, syn2",
          "when_to_prefer": "Default for country queries."
        }
      ]
    }
  ],
  "schema_exclusions": [
    {
      "overlap_group": "Discount",
      "ambiguity_class": "B – Granularity Levels",
      "table": "TableName",
      "column": "ColumnName",
      "reason": "Redundant sortable representation"
    }
  ]
}

Requires: pip install openpyxl
"""

from __future__ import annotations

import argparse
import hashlib
import json
import sys
import time
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Plugin root directory — all outputs go under <plugin_root>/output/
_PLUGIN_ROOT = Path(__file__).resolve().parent.parent.parent.parent
_DEFAULT_OUTPUT_DIR = _PLUGIN_ROOT / "output"

# ── Style constants ──────────────────────────────────────────────────────────
# Distinct header fill per sheet (skill spec: blue, teal, orange, purple)
BLUE_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
TEAL_FILL = PatternFill(start_color="00796B", end_color="00796B", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="BF5700", end_color="BF5700", fill_type="solid")
PURPLE_FILL = PatternFill(start_color="6A1B9A", end_color="6A1B9A", fill_type="solid")

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
MERGE_ALIGN = Alignment(wrap_text=True, vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
GUIDANCE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
GUIDANCE_FONT = Font(name="Calibri", size=10, italic=True, color="333333")
REVIEW_STATUS_OPTIONS = '"Pending,Approved,Rejected,Needs Discussion"'
CHECKLIST_STATUS_OPTIONS = '"Not Started,In Progress,Done,Blocked"'
AMBIGUITY_CLASS_OPTIONS = '"A – Competing Classifications,B – Granularity Levels,C – Multiple Representations,D – Cross-Table Duplicates,E – Context-Resolvable Variants"'
PRIORITY_OPTIONS = '"Critical,High,Medium,Low"'
EXCLUSION_SCOPE_OPTIONS = '"AI Data Schema Only,Model Level (isHidden)"'

# ── Sheet headers (4 sheets per skill spec) ──────────────────────────────────

# Sheet 1 — Column Descriptions (blue header)
COL_DESC_HEADERS = [
    "Overlap Group", "Ambiguity Class",
    "Table", "Column",
    "Current Description (read-only)",
    "Proposed Definition",
    "Suggested Synonyms",
    "Review Status", "Reviewer", "Reviewer Comments",
]

# Sheet 2 — Business Rules (teal header)
BIZ_RULES_HEADERS = [
    "Overlap Group", "Ambiguity Class",
    "Domain",
    "When User Says...",
    "Default Column", "Table",
    "When to Prefer Alternative",
    "Alternative Columns",
    "Priority",
    "Review Status", "Reviewer", "Reviewer Comments",
]

# Sheet 3 — Schema Exclusions (orange header)
EXCLUSION_HEADERS = [
    "Overlap Group", "Ambiguity Class",
    "Table", "Column", "Reason for Exclusion",
    "Exclusion Scope",
    "Action (Post-Approval)",
    "Review Status", "Reviewer", "Reviewer Comments",
]

# Sheet 4 — Implementation Checklist (purple header)
CHECKLIST_HEADERS = [
    "Step", "Action", "Source Sheet", "Details",
    "Status", "Owner", "Notes",
]

# 1-based column indices of group-level columns to merge (Sheet 1)
_COL_DESC_MERGE_COLS = [1, 2]

# ── Per-sheet guidance text (row 2, yellow background) ───────────────────────
_GUIDANCE = {
    "Column Descriptions": (
        "ACTION: Edit 'Proposed Definition' to describe what this column IS. "
        "'Current Description' is read-only (shows what's in the model today). "
        "The system will auto-add 'NOT competing-column' and 'Primary/Prefer X' "
        "based on Business Rules. Set Review Status → Approved when done."
    ),
    "Business Rules": (
        "ACTION: Review routing rules. 'Default Column' is the canonical column "
        "Copilot/agents will use when user queries match 'When User Says'. "
        "Change 'Default Column' to swap the primary — this auto-cascades to all "
        "column descriptions in the group. Set Review Status → Approved when done."
    ),
    "Schema Exclusions": (
        "ACTION: Review columns to hide from AI queries. Choose 'Exclusion Scope': "
        "'AI Data Schema Only' (safer, manual in Fabric UI — column stays in reports) "
        "or 'Model Level' (auto-applied via MCP — hides from reports AND AI). "
        "Set Review Status → Approved when done."
    ),
    "Implementation Checklist": (
        "INFO: Post-review execution steps. These run automatically after you approve "
        "items on the other sheets. Track progress here using the Status column."
    ),
}

# ── Standard implementation checklist (model-agnostic) ───────────────────────
CHECKLIST_DATA = [
    [1, "Hide attributes from AI data schema",
     "Schema Exclusions",
     "For every Approved row in Schema Exclusions, mark the column as hidden "
     "in Prep data for AI > AI data schema so Copilot/agents never see it.",
     "", "", ""],
    [2, "Update column descriptions in semantic model",
     "Column Descriptions",
     "For every Approved row, update the column's Description property in "
     "the semantic model. Synonyms from the 'Suggested Synonyms' column are "
     "automatically appended as 'Also known as: ...' in the description.",
     "", "", ""],
    [3, "Update AI instructions (Prep data for AI)",
     "Business Rules",
     "Generate AI instructions from approved Business Rules that have routing "
     "logic (When User Says + Default Column). Paste into Prep data for AI.",
     "", "", ""],
    [4, "Validate end-to-end",
     "All sheets",
     "Test with sample queries to confirm Copilot/agent picks the correct "
     "column for each ambiguous domain.",
     "", "", ""],
]


# ── Cell sanitization (prevent Excel formula injection) ──────────────────────

_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r", "\n")


def _sanitize_cell_value(value):
    """Prefix a leading apostrophe if a string value starts with a formula char.

    Excel interprets cell values starting with =, +, -, @, tab, CR, or LF as
    formulas. Since all data here is model metadata or sample values (never
    intentional formulas), we neutralize them with a leading apostrophe which
    Excel displays invisibly but prevents formula evaluation.
    """
    if isinstance(value, str) and value and value[0] in _FORMULA_PREFIXES:
        return "'" + value
    return value


# ── Row fingerprinting (for tamper detection at apply time) ──────────────────

def _row_fingerprint(*fields: str) -> str:
    """Compute a SHA-256 fingerprint from key identity fields."""
    canonical = "|".join(f.strip() for f in fields)
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()


def _generate_manifest(
    data: Dict[str, Any],
    analysis_rows: List[list],
    biz_rules_rows: List[list],
    excl_rows: List[list],
    excel_path: Path,
) -> Path:
    """Save a manifest JSON alongside the Excel for apply-time validation.

    The manifest contains SHA-256 fingerprints of each row's immutable
    identity fields so the apply script can detect tampered or user-added rows.
    """
    desc_fps = []
    for row in analysis_rows:
        # row layout: [overlap_group, ambiguity_class, table, column, ...]
        desc_fps.append(_row_fingerprint(
            "column_descriptions", row[0], row[2], row[3],
        ))

    rule_fps = []
    for row in biz_rules_rows:
        # row layout: [overlap_group, ambiguity_class, domain, ...]
        rule_fps.append(_row_fingerprint(
            "business_rules", row[0], row[2],
        ))

    excl_fps = []
    for row in excl_rows:
        # row layout: [overlap_group, ambiguity_class, table, column, ...]
        excl_fps.append(_row_fingerprint(
            "schema_exclusions", row[0], row[2], row[3],
        ))

    manifest = {
        "generated_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "model_name": data.get("model_name", ""),
        "workspace": data.get("workspace", ""),
        "excel_filename": excel_path.name,
        "row_fingerprints": {
            "column_descriptions": desc_fps,
            "schema_exclusions": excl_fps,
            "business_rules": rule_fps,
        },
    }

    manifest_path = excel_path.parent / "disambiguation_manifest.json"
    manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    return manifest_path


# ── Data conversion ──────────────────────────────────────────────────────────

def _overlap_groups_to_col_desc_rows(items: List[Dict[str, Any]]) -> List[list]:
    """Convert overlap_groups JSON items to Column Descriptions sheet rows.

    Only the "Is (Definition)" is stored per column. "Is Not" and "Canonical
    Note" are auto-composed at apply time from:
      - Business Rules sheet (which column is default → canonical note)
      - Overlap group membership (sibling columns → "Is Not")
    """
    rows = []
    for group in items:
        overlap_group = group.get("overlap_group", "")
        ambiguity_class = group.get("ambiguity_class", "")

        for col in group.get("columns", []):
            # Extract "Is" from proposed_description (first line)
            proposed = col.get("proposed_description", "")
            is_def = proposed.split("\n")[0].rstrip(".") + "." if proposed else ""

            rows.append([
                overlap_group,
                ambiguity_class,
                col.get("table", ""),
                col.get("column", ""),
                col.get("current_description", ""),
                is_def,
                col.get("suggested_synonyms", ""),
                "",  # Review Status
                "",  # Reviewer
                "",  # Reviewer Comments
            ])
    rows.sort(key=lambda r: r[0])
    return rows


def _overlap_groups_to_biz_rules_rows(items: List[Dict[str, Any]]) -> List[list]:
    """Convert overlap_groups JSON items to Business Rules sheet rows.

    One row per overlap group that has routing logic (non-empty
    when_user_says). Class E groups are excluded (no routing needed).
    """
    rows = []
    for group in items:
        ambiguity_class = group.get("ambiguity_class", "")
        when_user_says = group.get("when_user_says", "")
        default_column = group.get("default_column", "")

        # Skip Class E groups (context-resolvable, no rules needed)
        class_code = ambiguity_class.split("–")[0].split("—")[0].strip().rstrip(" \t")
        if class_code == "E":
            continue
        # Skip groups with no routing logic
        if not when_user_says and not default_column:
            continue

        # Collect alternative columns (columns that aren't the default)
        default_table = ""
        alt_cols = []
        when_to_prefer_parts = []
        for col in group.get("columns", []):
            full_name = f"{col.get('table', '')}.{col.get('column', '')}"
            if full_name == default_column:
                default_table = col.get("table", "")
            else:
                alt_cols.append(full_name)
                wtp = col.get("when_to_prefer", "")
                if wtp:
                    when_to_prefer_parts.append(wtp)

        # If default_column has table prefix, extract table
        if "." in default_column and not default_table:
            default_table = default_column.split(".")[0]

        rows.append([
            group.get("overlap_group", ""),
            ambiguity_class,
            group.get("domain", ""),
            when_user_says,
            default_column,
            default_table,
            "; ".join(when_to_prefer_parts) if when_to_prefer_parts else "",
            ", ".join(alt_cols) if alt_cols else "",
            group.get("priority", ""),
            "",  # Review Status
            "",  # Reviewer
            "",  # Reviewer Comments
        ])
    rows.sort(key=lambda r: r[0])
    return rows


def _exclusions_to_rows(items: List[Dict[str, Any]]) -> List[list]:
    """Convert schema_exclusions JSON items to sheet rows."""
    rows = []
    for item in items:
        rows.append([
            item.get("overlap_group", ""),
            item.get("ambiguity_class", ""),
            item.get("table", ""),
            item.get("column", ""),
            item.get("reason", ""),
            "AI Data Schema Only",  # Default to safer option
            "Hide from AI data schema",
            "",  # Review Status
            "",  # Reviewer
            "",  # Comments
        ])
    return rows


# ── Formatting helper ───────────────────────────────────────────────────────

def _style_sheet(ws, headers, data, col_widths, header_fill,
                 review_status_col=None, guidance_text=None):
    """Apply headers, guidance row, data, formatting, auto-filter, and dropdowns."""
    # Row 1: headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = HEADER_FONT
        cell.alignment = WRAP
        cell.border = THIN_BORDER

    # Row 2: guidance row (merged across all columns, yellow background)
    data_start_row = 2
    if guidance_text:
        data_start_row = 3
        ws.merge_cells(start_row=2, start_column=1,
                       end_row=2, end_column=len(headers))
        guide_cell = ws.cell(row=2, column=1, value=guidance_text)
        guide_cell.fill = GUIDANCE_FILL
        guide_cell.font = GUIDANCE_FONT
        guide_cell.alignment = Alignment(wrap_text=True, vertical="center")
        guide_cell.border = THIN_BORDER
        ws.row_dimensions[2].height = 36

    # Data rows
    for row_idx, row_data in enumerate(data, data_start_row):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx,
                          value=_sanitize_cell_value(value))
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = THIN_BORDER
            # Alternating row shading
            if row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL

    last_row = max(data_start_row, data_start_row + len(data) - 1)
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_row}"

    if review_status_col is not None and data:
        dv = DataValidation(
            type="list", formula1=REVIEW_STATUS_OPTIONS, allow_blank=True,
        )
        dv.error = "Select: Pending, Approved, Rejected, or Needs Discussion"
        dv.errorTitle = "Invalid status"
        col_letter = get_column_letter(review_status_col)
        dv.add(f"{col_letter}{data_start_row}:{col_letter}{last_row}")
        ws.add_data_validation(dv)

    # Add Ambiguity Class dropdown (find column dynamically)
    if data:
        for i, h in enumerate(headers):
            if h == "Ambiguity Class":
                col_letter = get_column_letter(i + 1)
                dv_class = DataValidation(
                    type="list", formula1=AMBIGUITY_CLASS_OPTIONS, allow_blank=True,
                )
                dv_class.error = "Select an ambiguity class (A–E)"
                dv_class.errorTitle = "Invalid class"
                dv_class.add(f"{col_letter}{data_start_row}:{col_letter}{last_row}")
                ws.add_data_validation(dv_class)
                break

    return data_start_row  # caller may need this for additional dropdowns


def _merge_group_cells(ws, num_data_rows, group_col_idx=1, merge_cols=None,
                       data_start_row=3):
    """Merge cells for group-level columns based on contiguous Overlap Group values.

    Parameters
    ----------
    ws : Worksheet
        The worksheet to merge cells in.
    num_data_rows : int
        Number of data rows (excluding header and guidance row).
    group_col_idx : int
        1-based column index used to detect group boundaries
        (default: 1 = Overlap Group).
    merge_cols : list[int] | None
        1-based column indices to merge. If None, uses [group_col_idx].
    data_start_row : int
        Row where data begins (after header + guidance row).
    """
    if num_data_rows < 2:
        return
    if merge_cols is None:
        merge_cols = [group_col_idx]

    current_val = ws.cell(row=data_start_row, column=group_col_idx).value
    group_start = data_start_row

    last_data_row = data_start_row + num_data_rows - 1
    for row_num in range(data_start_row + 1, last_data_row + 1):
        cell_val = ws.cell(row=row_num, column=group_col_idx).value
        if cell_val != current_val:
            if row_num - 1 > group_start:
                for col in merge_cols:
                    ws.merge_cells(
                        start_row=group_start, start_column=col,
                        end_row=row_num - 1, end_column=col,
                    )
                    ws.cell(row=group_start, column=col).alignment = MERGE_ALIGN
            current_val = cell_val
            group_start = row_num

    # Merge last group
    if last_data_row > group_start:
        for col in merge_cols:
            ws.merge_cells(
                start_row=group_start, start_column=col,
                end_row=last_data_row, end_column=col,
            )
            ws.cell(row=group_start, column=col).alignment = MERGE_ALIGN


# ── Public API ───────────────────────────────────────────────────────────────

def generate_workbook(
    data: Dict[str, Any],
    output_path: Path,
    *,
    fallback_names: bool = True,
) -> Path:
    """Generate the disambiguation review workbook from analysis data.

    Parameters
    ----------
    data : dict
        Analysis results with keys: overlap_groups, schema_exclusions.
        Optional: model_name, workspace.
    output_path : Path
        Where to save the .xlsx file.
    fallback_names : bool
        If True and output_path is locked, try _v2, _v3 suffixes.

    Returns
    -------
    Path
        The actual path where the file was saved.
    """
    analysis_rows = _overlap_groups_to_col_desc_rows(data.get("overlap_groups", []))
    biz_rules_rows = _overlap_groups_to_biz_rules_rows(data.get("overlap_groups", []))
    excl_rows = _exclusions_to_rows(data.get("schema_exclusions", []))

    wb = Workbook()

    # Sheet 1 — Column Descriptions (blue header)
    ws1 = wb.active
    ws1.title = "Column Descriptions"
    ds1 = _style_sheet(
        ws1, COL_DESC_HEADERS, analysis_rows,
        [22, 28, 24, 28, 40, 55, 35, 16, 14, 35],
        header_fill=BLUE_FILL,
        review_status_col=8,
        guidance_text=_GUIDANCE["Column Descriptions"],
    )
    # Merge group-level columns (Overlap Group, Ambiguity Class)
    _merge_group_cells(ws1, len(analysis_rows), group_col_idx=1,
                       merge_cols=_COL_DESC_MERGE_COLS, data_start_row=ds1)

    # Sheet 2 — Business Rules (teal header)
    ws2 = wb.create_sheet("Business Rules")
    ds2 = _style_sheet(
        ws2, BIZ_RULES_HEADERS, biz_rules_rows,
        [22, 28, 18, 42, 30, 24, 45, 40, 10, 16, 14, 35],
        header_fill=TEAL_FILL,
        review_status_col=10,
        guidance_text=_GUIDANCE["Business Rules"],
    )
    # Add Priority dropdown
    if biz_rules_rows:
        dv_priority = DataValidation(
            type="list", formula1=PRIORITY_OPTIONS, allow_blank=True,
        )
        dv_priority.error = "Select: Critical, High, Medium, or Low"
        dv_priority.errorTitle = "Invalid priority"
        last_br = ds2 + len(biz_rules_rows) - 1
        dv_priority.add(f"I{ds2}:I{last_br}")
        ws2.add_data_validation(dv_priority)

    # Sheet 3 — Schema Exclusions (orange header)
    ws3 = wb.create_sheet("Schema Exclusions")
    ds3 = _style_sheet(
        ws3, EXCLUSION_HEADERS, excl_rows,
        [20, 28, 24, 35, 55, 26, 24, 16, 14, 35],
        header_fill=ORANGE_FILL,
        review_status_col=8,
        guidance_text=_GUIDANCE["Schema Exclusions"],
    )
    # Add Exclusion Scope dropdown (column 6 = F)
    if excl_rows:
        dv_scope = DataValidation(
            type="list", formula1=EXCLUSION_SCOPE_OPTIONS, allow_blank=True,
        )
        dv_scope.error = "Select: AI Data Schema Only or Model Level (isHidden)"
        dv_scope.errorTitle = "Invalid scope"
        last_se = ds3 + len(excl_rows) - 1
        dv_scope.add(f"F{ds3}:F{last_se}")
        ws3.add_data_validation(dv_scope)

    # Sheet 4 — Implementation Checklist (purple header)
    ws4 = wb.create_sheet("Implementation Checklist")
    ds4 = _style_sheet(
        ws4, CHECKLIST_HEADERS, CHECKLIST_DATA,
        [6, 38, 20, 70, 14, 14, 35],
        header_fill=PURPLE_FILL,
        review_status_col=5,
        guidance_text=_GUIDANCE["Implementation Checklist"],
    )
    # Override the Status dropdown with checklist-specific options
    ws4.data_validations.dataValidation.clear()
    dv = DataValidation(
        type="list", formula1=CHECKLIST_STATUS_OPTIONS, allow_blank=True,
    )
    dv.error = "Select: Not Started, In Progress, Done, or Blocked"
    dv.errorTitle = "Invalid status"
    dv.add(f"E{ds4}:E{ds4 + len(CHECKLIST_DATA) - 1}")
    ws4.add_data_validation(dv)

    # Save — try primary, then fallback names if file is locked
    output_path.parent.mkdir(parents=True, exist_ok=True)
    candidates = [output_path]
    if fallback_names:
        stem = output_path.stem
        suffix = output_path.suffix
        parent = output_path.parent
        candidates.extend([
            parent / f"{stem}_v2{suffix}",
            parent / f"{stem}_v3{suffix}",
        ])

    saved_path = None
    for candidate in candidates:
        try:
            wb.save(candidate)
            saved_path = candidate
            break
        except PermissionError:
            continue

    if saved_path is None:
        raise PermissionError(
            f"All output filenames are locked. Close Excel and retry. "
            f"Tried: {', '.join(str(c) for c in candidates)}"
        )

    print(f"Saved: {saved_path}")
    print(f"  Column Descriptions:      {len(analysis_rows)} rows")
    print(f"  Business Rules:           {len(biz_rules_rows)} rows")
    print(f"  Schema Exclusions:        {len(excl_rows)} rows")
    print(f"  Implementation Checklist: {len(CHECKLIST_DATA)} steps")

    # Generate tamper-detection manifest alongside the workbook
    manifest_path = _generate_manifest(
        data, analysis_rows, biz_rules_rows, excl_rows, saved_path,
    )
    print(f"  Manifest:                 {manifest_path}")

    return saved_path


def main():
    parser = argparse.ArgumentParser(
        description="Generate disambiguation review workbook from analysis JSON"
    )
    parser.add_argument(
        "--input", "-i", type=Path, default=None,
        help="Path to analysis JSON file. If omitted, reads from stdin.",
    )
    parser.add_argument(
        "--output", "-o", type=Path, default=_DEFAULT_OUTPUT_DIR / "disambiguation_results.xlsx",
        help="Output Excel file path (default: <plugin>/output/disambiguation_results.xlsx)",
    )
    args = parser.parse_args()

    # Read JSON input
    if args.input:
        data = json.loads(args.input.read_text(encoding="utf-8-sig"))
    else:
        data = json.load(sys.stdin)

    generate_workbook(data, args.output)


if __name__ == "__main__":
    main()
